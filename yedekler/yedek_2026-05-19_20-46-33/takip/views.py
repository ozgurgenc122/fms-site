from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
import os
import zipfile
from io import BytesIO
from django.conf import settings
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from decimal import Decimal
from .models import Is, RiskKaydi


DATE_FIELDS = [
    "planlanan_yukleme_zamani",
    "gerceklesen_yukleme_zamani",
    "bosaltma_evrak_teslim_zamani",
    "bosaltma_zamani",
    "ikinci_planlanan_yukleme_zamani",
    "ikinci_gerceklesen_yukleme_zamani",
    "ikinci_bosaltma_evrak_teslim_zamani",
    "ikinci_bosaltma_zamani",
    "gemi_kalkis_tarihi",
    "gemi_yanasma_tarihi",
    "limandan_cikis_tarihi",
]

INT_FIELDS = [
    "yukleme_ucretsiz_saat",
    "bosaltma_ucretsiz_saat",
    "ikinci_yukleme_ucretsiz_saat",
    "ikinci_bosaltma_ucretsiz_saat",
    "liman_ucretsiz_saat",
]

DECIMAL_FIELDS = [
    "derece_min",
    "derece_max",
    "yukleme_bekleme_diger_tutar",
    "bosaltma_bekleme_diger_tutar",
    "ikinci_yukleme_bekleme_diger_tutar",
    "ikinci_bosaltma_bekleme_diger_tutar",
    "liman_bekleme_diger_tutar",
    "elektronka_bedeli",
    "liman_masrafi",
    "aktarma_masrafi",
    "dozvola_masrafi",
    "tasima1_gelen_fatura_tutari",
    "tasima1_elden_verilen_avans",
    "tasima1_yapilan_odeme",
    "tasima2_gelen_fatura_tutari",
    "tasima2_elden_verilen_avans",
    "tasima2_yapilan_odeme",
    "gelen_fatura_tutari",
    "fatura_tutari",
    "hesaplanan_teminat",
    "fatura_edilecek_teminat_tutari",
    "satis_faturasi_tutari",
    "gemi_faturasi_tutari",
]

FILE_FIELDS = [
    "yukleme_evraklari",
    "gumruk_evraklari",
    "yurtdisi_evraklari",
    "bosaltma_evraklari",
]


def temiz_decimal(deger):
    if deger in [None, ""]:
        return Decimal("0")
    try:
        return Decimal(str(deger).replace(",", "."))
    except Exception:
        return Decimal("0")


def temiz_int(deger):
    if deger in [None, ""]:
        return None
    try:
        return int(deger)
    except Exception:
        return None


def yeniden_sirala():
    for index, kayit in enumerate(Is.objects.all().order_by("id"), start=1):
        if kayit.sira_no != index:
            kayit.sira_no = index
            kayit.save(update_fields=["sira_no"])


def evrak_sil(kayit, alan):
    eski_dosya = getattr(kayit, alan, None)
    if eski_dosya:
        try:
            eski_dosya.delete(save=False)
        except Exception:
            pass
    setattr(kayit, alan, None)


def risk_uyarilari_bul(kayit=None, sofor_adi="", sofor_no="", nakliyeci="", plaka=""):
    sorgu = Is.objects.all()
    if kayit and kayit.id:
        sorgu = sorgu.exclude(id=kayit.id)

    uyarilar = []

    if nakliyeci:
        for i in sorgu.filter(nakliyeci__icontains=nakliyeci).exclude(nakliyeci_notu="")[:5]:
            uyarilar.append(f"⚠️ Nakliyeci uyarısı: {i.nakliyeci} / {i.nakliyeci_notu}")

    if sofor_adi:
        for i in sorgu.filter(sofor_adi__icontains=sofor_adi).exclude(sofor_notu="")[:5]:
            uyarilar.append(f"⚠️ Şoför uyarısı: {i.sofor_adi} / {i.sofor_notu}")

    if sofor_no:
        for i in sorgu.filter(sofor_no__icontains=sofor_no).exclude(sofor_notu="")[:5]:
            uyarilar.append(f"⚠️ Şoför telefon uyarısı: {i.sofor_no} / {i.sofor_notu}")

    if plaka:
        for i in sorgu.filter(plaka__icontains=plaka).exclude(risk_notu="")[:5]:
            uyarilar.append(f"⚠️ Plaka uyarısı: {i.plaka} / {i.risk_notu}")

    return uyarilar


def kaydi_formdan_doldur(kayit, request, yeni_kayit=False):
    for alan, deger in request.POST.items():
        if alan == "csrfmiddlewaretoken":
            continue
        if alan.endswith("_sil"):
            continue
        if alan == "kontrol":
            continue
        if not hasattr(kayit, alan):
            continue

        if alan in DATE_FIELDS:
            if deger:
                setattr(kayit, alan, deger)
            elif yeni_kayit:
                setattr(kayit, alan, None)

        elif alan in INT_FIELDS:
            temiz = temiz_int(deger)
            if temiz is not None:
                setattr(kayit, alan, temiz)

        elif alan in DECIMAL_FIELDS:
            setattr(kayit, alan, temiz_decimal(deger))

        else:
            if yeni_kayit or deger != "":
                setattr(kayit, alan, deger)

    for alan in FILE_FIELDS:
        if request.POST.get(f"{alan}_sil") == "1":
            evrak_sil(kayit, alan)

        if request.FILES.get(alan):
            evrak_sil(kayit, alan)
            setattr(kayit, alan, request.FILES.get(alan))

    kayit.save()
    return kayit


@login_required(login_url="login")
def ana_sayfa(request):
    yeniden_sirala()

    isler = Is.objects.exclude(durum="Tamamlandı").order_by(
        "planlanan_yukleme_zamani",
        "gerceklesen_yukleme_zamani",
        "sira_no",
        "id"
    )

    baslangic = request.GET.get("baslangic", "")
    bitis = request.GET.get("bitis", "")
    musteri = request.GET.get("musteri", "")
    durum = request.GET.get("durum", "")
    odeme = request.GET.get("odeme", "")
    plaka = request.GET.get("plaka", "")
    hat = request.GET.get("hat", "")

    if baslangic:
        isler = isler.filter(olusturma_tarihi_date_gte=baslangic)
    if bitis:
        isler = isler.filter(olusturma_tarihi_date_lte=bitis)
    if musteri:
        isler = isler.filter(musteri_adi__icontains=musteri)
    if durum:
        isler = isler.filter(durum=durum)
    if odeme:
        isler = isler.filter(odeme_durumu=odeme)
    if plaka:
        isler = isler.filter(plaka__icontains=plaka)
    if hat:
        isler = isler.filter(hat__icontains=hat)

    devam_eden_isler = Is.objects.exclude(
        durum__in=["Tamamlandı", "İptal"]
    ).order_by("sira_no", "id")

    toplamlar = {
        "EUR": {"maliyet": Decimal("0"), "satis": Decimal("0"), "kar": Decimal("0")},
        "USD": {"maliyet": Decimal("0"), "satis": Decimal("0"), "kar": Decimal("0")},
        "TRY": {"maliyet": Decimal("0"), "satis": Decimal("0"), "kar": Decimal("0")},
    }

    for i in isler:
        pb = i.para_birimi or "EUR"

        if pb not in toplamlar:
            continue

        toplamlar[pb]["maliyet"] += i.toplam_maliyet()
        toplamlar[pb]["satis"] += i.toplam_satis()
        toplamlar[pb]["kar"] += i.net_kar()

    musteriler = {}
    for kayit in isler:
        musteri_adi = kayit.musteri_adi or "Müşteri Girilmemiş"
        musteriler.setdefault(musteri_adi, []).append(kayit)

    return render(request, "takip/ana_sayfa.html", {
        "isler": isler,
        "devam_eden_isler": devam_eden_isler,
        "musteriler": musteriler,
        "toplamlar": toplamlar,
        "baslangic": baslangic,
        "bitis": bitis,
        "musteri": musteri,
        "durum": durum,
        "odeme": odeme,
        "plaka": plaka,
        "hat": hat,
    })


@login_required(login_url="login")
def is_ekle(request):
    uyarilar = []

    if request.method == "POST":
        kontrol = request.POST.get("kontrol", "")

        uyarilar = risk_uyarilari_bul(
            sofor_adi=request.POST.get("sofor_adi", ""),
            sofor_no=request.POST.get("sofor_no", ""),
            nakliyeci=request.POST.get("nakliyeci", ""),
            plaka=request.POST.get("plaka", ""),
        )

        if kontrol == "1" and uyarilar:
            return render(request, "takip/ekle.html", {
                "uyarilar": uyarilar,
                "post": request.POST,
            })

        kayit = Is.objects.create(sira_no=Is.objects.count() + 1)
        kaydi_formdan_doldur(kayit, request, yeni_kayit=True)
        yeniden_sirala()
        return redirect("ana_sayfa")

    return render(request, "takip/ekle.html", {"uyarilar": uyarilar})


@login_required(login_url="login")
def is_duzenle(request, id):
    kayit = get_object_or_404(Is, id=id)
    uyarilar = []

    if request.method == "POST":

        if request.POST.get("durum") == "Tamamlandı":
            if not kayit.bosaltma_evraklari and not request.FILES.get("bosaltma_evraklari"):
                uyarilar.append("⚠️ Boşaltma evrakı yüklenmeden iş tamamlanamaz.")
                return render(request, "takip/duzenle.html", {
                    "is": kayit,
                    "uyarilar": uyarilar,
                })

        kaydi_formdan_doldur(kayit, request, yeni_kayit=False)
        yeniden_sirala()
        return redirect("ana_sayfa")

    return render(request, "takip/duzenle.html", {
        "is": kayit,
        "uyarilar": uyarilar,
    })


@login_required(login_url="login")
def is_goster(request, id):
    kayit = get_object_or_404(Is, id=id)
    return render(request, "takip/goster.html", {"is": kayit})


@login_required(login_url="login")
def is_sil(request, id):
    kayit = get_object_or_404(Is, id=id)
    kayit.delete()
    yeniden_sirala()
    return redirect("ana_sayfa")


@login_required(login_url="login")
def is_kopyala(request, id):
    eski = get_object_or_404(Is, id=id)

    Is.objects.create(
        sira_no=Is.objects.count() + 1,
        musteri_adi=eski.musteri_adi,
        isin_adi="",
        durum="Teklif",
        odeme_durumu="Ödenmedi",
        hat=eski.hat,
        para_birimi=eski.para_birimi,
        derece_min=eski.derece_min,
        derece_max=eski.derece_max,
        arac_tipi=eski.arac_tipi,
        yukleme_konum=eski.yukleme_konum,
        bosaltma_konum=eski.bosaltma_konum,
        ikinci_tasima_var_mi=eski.ikinci_tasima_var_mi,
        ikinci_arac_tipi=eski.ikinci_arac_tipi,
        ikinci_yukleme_konum=eski.ikinci_yukleme_konum,
        ikinci_bosaltma_konum=eski.ikinci_bosaltma_konum,
        gemi_adi=eski.gemi_adi,
    )

    yeniden_sirala()
    return redirect("ana_sayfa")


@login_required(login_url="login")
def aktif_ise_gecir(request, id):
    kayit = get_object_or_404(Is, id=id)

    if kayit.durum == "Teklif":
        kayit.durum = "Planlandı"
        kayit.save()

    return redirect("ana_sayfa")


def excel_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url="login")
def excel_sablon(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "İş Import Şablonu"

    ws.append([
        "musteri_adi", "isin_adi", "hat", "durum", "odeme_durumu", "para_birimi",
        "derece_min", "derece_max", "arac_tipi", "plaka", "sofor_adi", "sofor_no",
        "nakliyeci", "guncel_konum",
        "tasima1_gelen_fatura", "tasima1_gelen_fatura_tutari",
        "tasima1_elden_verilen_avans", "tasima1_yapilan_odeme",
        "ikinci_tasima_var_mi", "ikinci_arac_tipi", "ikinci_plaka",
        "ikinci_sofor_adi", "ikinci_sofor_no", "ikinci_nakliyeci", "ikinci_guncel_konum",
        "tasima2_gelen_fatura", "tasima2_gelen_fatura_tutari",
        "tasima2_elden_verilen_avans", "tasima2_yapilan_odeme",
        "gemi_adi", "elektronka_bedeli", "liman_masrafi", "aktarma_masrafi", "dozvola_masrafi",
        "gelen_fatura_numarasi", "gelen_fatura_tutari",
        "esya_cinsi", "kap", "kg", "gtip", "fatura_tutari",
        "hesaplanan_teminat", "fatura_edilecek_teminat_tutari",
        "teminat_satisa_dahil_mi", "teminat_ayri_fatura_no",
        "satis_faturasi", "satis_faturasi_tutari",
        "gelen_gemi_faturasi", "gemi_faturasi_tutari",
    ])

    return excel_response(wb, "is_import_sablonu.xlsx")


@login_required(login_url="login")
def excel_import(request):
    mesaj = ""

    if request.method == "POST":
        dosya = request.FILES.get("excel_dosyasi")

        if not dosya:
            mesaj = "Excel dosyası seçilmedi."
        else:
            wb = load_workbook(dosya)
            ws = wb.active
            basliklar = [cell.value for cell in ws[1]]
            sayac = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                veri = dict(zip(basliklar, row))

                if not veri.get("musteri_adi") and not veri.get("isin_adi"):
                    continue

                kayit = Is.objects.create(sira_no=Is.objects.count() + 1)

                for alan, deger in veri.items():
                    if not alan or not hasattr(kayit, alan):
                        continue

                    if alan in DECIMAL_FIELDS:
                        setattr(kayit, alan, temiz_decimal(deger))
                    elif alan in INT_FIELDS:
                        temiz = temiz_int(deger)
                        if temiz is not None:
                            setattr(kayit, alan, temiz)
                    elif deger is not None:
                        setattr(kayit, alan, deger)

                kayit.save()
                sayac += 1

            yeniden_sirala()
            mesaj = f"{sayac} adet iş başarıyla içe aktarıldı."

    return render(request, "takip/excel_import.html", {"mesaj": mesaj})


@login_required(login_url="login")
def excel_musteriler(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteri Raporu"

    ws.append(["Müşteri", "İş Sayısı", "Plaka", "Yükleme", "Boşaltma", "Para Birimi", "Toplam Kar"])

    musteriler = {}

    for kayit in Is.objects.all():
        m = kayit.musteri_adi or "Müşteri Girilmemiş"

        if m not in musteriler:
            musteriler[m] = {
                "sayi": 0,
                "plakalar": [],
                "yuklemeler": [],
                "bosaltmalar": [],
                "para_birimleri": [],
                "kar": Decimal("0"),
            }

        musteriler[m]["sayi"] += 1

        if kayit.plaka:
            musteriler[m]["plakalar"].append(kayit.plaka)
        if kayit.planlanan_yukleme_zamani:
            musteriler[m]["yuklemeler"].append(str(kayit.planlanan_yukleme_zamani))
        if kayit.bosaltma_zamani:
            musteriler[m]["bosaltmalar"].append(str(kayit.bosaltma_zamani))
        if kayit.para_birimi:
            musteriler[m]["para_birimleri"].append(kayit.para_birimi)

        musteriler[m]["kar"] += kayit.net_kar()

    for m, v in musteriler.items():
        ws.append([
            m,
            v["sayi"],
            ", ".join(v["plakalar"]),
            ", ".join(v["yuklemeler"]),
            ", ".join(v["bosaltmalar"]),
            ", ".join(sorted(set(v["para_birimleri"]))),
            v["kar"],
        ])

    return excel_response(wb, "musteri_raporu.xlsx")


@login_required(login_url="login")
def excel_isler(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tüm İşler"

    ws.append([
        "Sıra", "Müşteri", "İşin Adı", "Durum", "Ödeme", "Hat", "Para Birimi",
        "Derece Min", "Derece Max", "Araç Tipi", "Plaka", "Şoför", "Şoför No",
        "Nakliyeci", "Güncel Konum",
        "1. Taşıma Fatura", "1. Taşıma Fatura Tutarı", "1. Elden Avans",
        "1. Yapılan Ödeme", "1. Kalan Ödeme",
        "2. Taşıma Var mı", "2. Araç", "2. Plaka", "2. Şoför", "2. Nakliyeci",
        "2. Taşıma Fatura", "2. Taşıma Fatura Tutarı", "2. Elden Avans",
        "2. Yapılan Ödeme", "2. Kalan Ödeme",
        "Gemi", "Elektronka", "Liman", "Aktarma", "Dozvola",
        "Gelen Fatura", "Gelen Fatura Tutarı", "Eşya", "Kap", "Kg", "GTİP",
        "Fatura Tutarı", "Hesaplanan Teminat", "Fatura Edilecek Teminat",
        "Satış Faturası", "Satış Tutarı", "Gemi Faturası", "Gemi Faturası Tutarı",
        "Evrak Durumu", "Toplam Bekleme", "Toplam Maliyet", "Toplam Satış", "Net Kar",
    ])

    for i in Is.objects.all().order_by("sira_no"):
        ws.append([
            i.sira_no, i.musteri_adi, i.isin_adi, i.durum, i.odeme_durumu, i.hat, i.para_birimi,
            i.derece_min, i.derece_max, i.arac_tipi, i.plaka, i.sofor_adi, i.sofor_no,
            i.nakliyeci, i.guncel_konum,
            i.tasima1_gelen_fatura, i.tasima1_gelen_fatura_tutari,
            i.tasima1_elden_verilen_avans, i.tasima1_yapilan_odeme, i.tasima1_kalan_odeme(),
            i.ikinci_tasima_var_mi, i.ikinci_arac_tipi, i.ikinci_plaka, i.ikinci_sofor_adi, i.ikinci_nakliyeci,
            i.tasima2_gelen_fatura, i.tasima2_gelen_fatura_tutari,
            i.tasima2_elden_verilen_avans, i.tasima2_yapilan_odeme, i.tasima2_kalan_odeme(),
            i.gemi_adi, i.elektronka_bedeli, i.liman_masrafi, i.aktarma_masrafi, i.dozvola_masrafi,
            i.gelen_fatura_numarasi, i.gelen_fatura_tutari,
            i.esya_cinsi, i.kap, i.kg, i.gtip, i.fatura_tutari, i.hesaplanan_teminat,
            i.fatura_edilecek_teminat_tutari,
            i.satis_faturasi, i.satis_faturasi_tutari,
            i.gelen_gemi_faturasi, i.gemi_faturasi_tutari,
            i.evrak_durumu(), i.toplam_bekleme_tutari(), i.toplam_maliyet(), i.toplam_satis(), i.net_kar(),
        ])

    return excel_response(wb, "tum_isler.xlsx")


@login_required(login_url="login")
def excel_konumlar(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Devam Eden Konumlar"

    ws.append(["Müşteri", "İşin Adı", "Plaka", "Güncel Konum", "Evrak Durumu"])

    isler = Is.objects.exclude(durum__in=["Tamamlandı", "İptal"])

    for i in isler:
        ws.append([i.musteri_adi, i.isin_adi, i.plaka, i.guncel_konum, i.evrak_durumu()])

        if i.ikinci_tasima_var_mi == "Evet":
            ws.append([i.musteri_adi, i.isin_adi, i.ikinci_plaka, i.ikinci_guncel_konum, i.evrak_durumu()])

    return excel_response(wb, "devam_eden_konumlar.xlsx")


@login_required(login_url="login")
def yedek_al(request):
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        db_path = settings.BASE_DIR / "db.sqlite3"

        if os.path.exists(db_path):
            zip_file.write(db_path, "db.sqlite3")

        media_root = settings.MEDIA_ROOT

        if os.path.exists(media_root):
            for klasor_yolu, klasorler, dosyalar in os.walk(media_root):
                for dosya in dosyalar:
                    tam_yol = os.path.join(klasor_yolu, dosya)
                    zip_icindeki_yol = os.path.relpath(tam_yol, settings.BASE_DIR)
                    zip_file.write(tam_yol, zip_icindeki_yol)

    zip_buffer.seek(0)

    tarih = datetime.now().strftime("%Y-%m-%d_%H-%M")
    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="fms_yedek_{tarih}.zip"'
    return response


@login_required(login_url="login")
def risk_listesi(request):
    riskler = RiskKaydi.objects.all().order_by("-tarih", "-id")
    return render(request, "takip/risk_listesi.html", {"riskler": riskler})


@login_required(login_url="login")
def risk_ekle(request):
    if request.method == "POST":
        RiskKaydi.objects.create(
            firma_adi=request.POST.get("firma_adi", ""),
            sofor_adi=request.POST.get("sofor_adi", ""),
            plaka=request.POST.get("plaka", ""),
            sorun_tipi=request.POST.get("sorun_tipi", "Diğer"),
            aciklama=request.POST.get("aciklama", ""),
        )
        return redirect("risk_listesi")

    return render(request, "takip/risk_ekle.html")


@login_required(login_url="login")
def risk_sil(request, id):
    kayit = get_object_or_404(RiskKaydi, id=id)
    kayit.delete()
    return redirect("risk_listesi")

@login_required(login_url="login")
def guvenli_cikis(request):
    logout(request)
    return redirect("login")

@login_required(login_url="login")
def tamamlanan_isler(request):
    isler = Is.objects.filter(durum="Tamamlandı").order_by(
        "planlanan_yukleme_zamani",
        "gerceklesen_yukleme_zamani",
        "sira_no",
        "id"
    )

    toplamlar = {
        "EUR": {
            "maliyet": Decimal("0"),
            "satis": Decimal("0"),
            "kar": Decimal("0"),
            "adet": 0,
        },
        "USD": {
            "maliyet": Decimal("0"),
            "satis": Decimal("0"),
            "kar": Decimal("0"),
            "adet": 0,
        },
        "TRY": {
            "maliyet": Decimal("0"),
            "satis": Decimal("0"),
            "kar": Decimal("0"),
            "adet": 0,
        },
    }

    for i in isler:
        pb = i.para_birimi or "EUR"

        if pb not in toplamlar:
            toplamlar[pb] = {
                "maliyet": Decimal("0"),
                "satis": Decimal("0"),
                "kar": Decimal("0"),
                "adet": 0,
            }

        toplamlar[pb]["maliyet"] += i.toplam_maliyet()
        toplamlar[pb]["satis"] += i.toplam_satis()
        toplamlar[pb]["kar"] += i.net_kar()
        toplamlar[pb]["adet"] += 1

    return render(request, "takip/tamamlanan_isler.html", {
        "isler": isler,
        "toplamlar": toplamlar,
    })

@login_required(login_url="login")
def cari_listesi(request):

    musteriler = {}

    for i in Is.objects.all():

        musteri = i.musteri_adi or "Müşteri Girilmemiş"

        if musteri not in musteriler:
            musteriler[musteri] = {
                "is_sayisi": 0,

                "EUR": {
                    "maliyet": Decimal("0"),
                    "satis": Decimal("0"),
                    "kar": Decimal("0"),
                },

                "USD": {
                    "maliyet": Decimal("0"),
                    "satis": Decimal("0"),
                    "kar": Decimal("0"),
                },

                "TRY": {
                    "maliyet": Decimal("0"),
                    "satis": Decimal("0"),
                    "kar": Decimal("0"),
                },
            }

        pb = i.para_birimi or "EUR"

        if pb not in musteriler[musteri]:
            musteriler[musteri][pb] = {
                "maliyet": Decimal("0"),
                "satis": Decimal("0"),
                "kar": Decimal("0"),
            }

        musteriler[musteri]["is_sayisi"] += 1

        musteriler[musteri][pb]["maliyet"] += i.toplam_maliyet()
        musteriler[musteri][pb]["satis"] += i.toplam_satis()
        musteriler[musteri][pb]["kar"] += i.net_kar()

    return render(request, "takip/cari_listesi.html", {
        "musteriler": musteriler
    })